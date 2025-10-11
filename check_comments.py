import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('c:/Users/ADMIN/Python/Sales_dashboard/Bomba Profitability/Bomba Foods-MIS.xlsx')
ws = wb['P&L (Niko)']

# Check for comments in the worksheet
comments_found = []
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.comment:
            comments_found.append({
                'row': row,
                'col': col,
                'comment': cell.comment.text,
                'value': cell.value
            })

print(f'Found {len(comments_found)} cells with comments')
for comment in comments_found[:5]:  # Show first 5 comments
    print(f'Cell {comment["row"]},{comment["col"]}: {comment["value"]} -> {comment["comment"][:50]}...')

# Also check sheet dimensions
print(f'Worksheet max row: {ws.max_row}')
print(f'Worksheet max column: {ws.max_column}')
