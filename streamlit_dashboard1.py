import streamlit as st
import pandas as pd
import os
import plotly.express as px
import numpy as np
import openpyxl
import json
from pathlib import Path
from bs4 import BeautifulSoup

st.set_page_config(page_title="Niko Foods Profitability Dashboard", layout="wide")

# File to store remarks
REMARKS_FILE = os.path.join(os.path.dirname(__file__), "dashboard_remarks.json")

# Load remarks from file
def load_remarks():
    if os.path.exists(REMARKS_FILE):
        try:
            with open(REMARKS_FILE, 'r') as f:
                return json.load(f)
        except Exception:
            return {"remarks": {}}
    return {"remarks": {}}

# Save remarks to file
def save_remarks(data):
    try:
        with open(REMARKS_FILE, 'w') as f:
            json.dump(data, f, indent=2)
        return True
    except Exception as e:
        st.error(f"Error saving remarks: {e}")
        return False

# Initialize session state for remarks
if 'remarks_data' not in st.session_state:
    st.session_state.remarks_data = load_remarks()
if 'show_remarks_editor' not in st.session_state:
    st.session_state.show_remarks_editor = False

# Inject CSS for sticky header, scrollable table container, and button styling
st.markdown(
    '''
    <style>
    .freeze-header-table-container {
        max-height: 600px;
        overflow-y: auto;
    }
    .freeze-header-table-container table {
        width: 100%;
        border-collapse: collapse;
    }
    thead th {
        position: sticky !important;
        top: 0;
        z-index: 2;
        background: #003366 !important;
        color: white !important;
    }
    /* Ensure remark highlighting takes priority */
    .freeze-header-table-container td[style*="#FFE4B5"] {
        background-color: #FFE4B5 !important;
        border: 3px solid #ff6b6b !important;
        font-weight: bold !important;
    }
    /* Enhanced download button styling */
    .stDownloadButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 12px 24px !important;
        font-weight: 600 !important;
        font-size: 16px !important;
        box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4) !important;
        transition: all 0.3s ease !important;
    }
    .stDownloadButton > button:hover {
        background: linear-gradient(135deg, #764ba2 0%, #667eea 100%) !important;
        box-shadow: 0 6px 20px rgba(102, 126, 234, 0.6) !important;
        transform: translateY(-2px) !important;
    }
    .stDownloadButton > button:active {
        transform: translateY(0px) !important;
    }
    </style>
    ''',
    unsafe_allow_html=True
)

# Load only the Niko sheet from Excel
file_path = os.path.join(os.path.dirname(__file__), "Bomba Foods-MIS.xlsx")
if not os.path.exists(file_path):
    st.error(f"Excel file '{file_path}' not found in this directory.")
    st.stop()

# Only load the Niko sheet
sheets_data = {'P&L (Niko)': pd.read_excel(file_path, sheet_name='P&L (Niko)')}

# Convert 'Month' column to string for all dataframes if it is datetime
def ensure_month_str(df):
    if 'Month' in df.columns and pd.api.types.is_datetime64_any_dtype(df['Month']):
        df['Month'] = df['Month'].dt.strftime('%Y-%m')
    return df

branch_frames = []
sheets_data_str = {}
for branch, df in sheets_data.items():
    df = ensure_month_str(df.copy())
    df['Branch'] = branch
    branch_frames.append(df)
    sheets_data_str[branch] = df

full_df = pd.concat(branch_frames, ignore_index=True)
full_df = ensure_month_str(full_df)

# Only show Niko branch
branch_names = ['P&L (Niko)']
branch_option = 'P&L (Niko)'

# Determine unhidden columns for the selected branch using openpyxl
import openpyxl
excel_path = file_path
wb = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
ws = wb[branch_option]

# Get indexes of unhidden columns (1-based for openpyxl)
unhidden_col_indexes = [
    idx for idx, col in enumerate(ws.iter_cols(1, ws.max_column), 1)
    if not ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].hidden
]

# Get indexes of unhidden rows (1-based for openpyxl, skip header row)
unhidden_row_indexes = [
    idx - 2 for idx in range(2, ws.max_row + 1)  # -2 to convert to 0-based pandas index (row 2 in Excel = index 0 in pandas)
    if not ws.row_dimensions[idx].hidden
]

# Map to DataFrame column names (0-based for pandas)
df_cols = list(sheets_data_str[branch_option].columns)
unhidden_cols = [df_cols[idx-1] for idx in unhidden_col_indexes if idx-1 < len(df_cols)]

# Filter both columns and rows
df_to_show = sheets_data_str[branch_option][unhidden_cols].copy()
if unhidden_row_indexes:
    df_to_show = df_to_show.iloc[unhidden_row_indexes].reset_index(drop=True)

# Download Excel button (only visible/unhidden columns)
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
# Write to Excel
towrite = io.BytesIO()
df_to_show.to_excel(towrite, index=False, engine='openpyxl')
towrite.seek(0)
wb = load_workbook(towrite)
ws = wb.active
# Header formatting
header_fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=14)
# Month column formatting
from datetime import datetime
month_fmt = '%b-%y'
for cell in ws[1]:
    # Try to parse as date for month columns
    try:
        # If header looks like a date, format as 'Apr-25'
        if isinstance(cell.value, str):
            try:
                dt = pd.to_datetime(cell.value)
                cell.value = dt.strftime(month_fmt)
            except Exception:
                pass
        elif isinstance(cell.value, datetime):
            cell.value = cell.value.strftime(month_fmt)
    except Exception:
        pass
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
# If any month columns have date values, format them too
for c in range(1, ws.max_column+1):
    col_val = ws.cell(row=1, column=c).value
    if col_val:
        try:
            dt = pd.to_datetime(col_val, errors='coerce')
            if not pd.isnull(dt):
                # Format all values in this column if they are dates
                for r in range(2, ws.max_row+1):
                    v = ws.cell(row=r, column=c).value
                    if isinstance(v, datetime):
                        ws.cell(row=r, column=c).number_format = 'mmm-yy'
        except Exception:
            pass
# Border for all cells
thin = Side(border_style="thin", color="000000")
for row in ws.iter_rows():
    for cell in row:
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
# Row/column/number formatting logic
particulars_col = None
for idx, cell in enumerate(ws[1], 1):
    if str(cell.value).strip().lower() == 'particulars':
        particulars_col = idx
        break
if particulars_col:
    sales_block = False
    blue_block = False
    green1_block = False
    green2_block = False
    red_block = False
    for r in range(2, ws.max_row+1):
        val = ws.cell(row=r, column=particulars_col).value
        style = None
        if isinstance(val, str):
            txt = val.strip().lower()
            
            # Block coloring logic (apply first, so specific rows can override)
            # Sales block: FOOD SALES to TOTAL SALES AND SERVICE CHARGES
            if txt in ['food sales', 'drinks sales', 'service charge', 'service charge ']:
                sales_block = True
            if sales_block:
                style = {'fill': PatternFill(start_color='fff9c4', end_color='fff9c4', fill_type='solid')}
                if txt == 'total sales and service charges':
                    style = {'fill': PatternFill(start_color='ffe066', end_color='ffe066', fill_type='solid'), 'font': Font(bold=True)}
                    sales_block = False
            
            pink_rows = ['less: discount', 'less: adjusted ( net of gst)', 'net discount']
            if txt in pink_rows:
                style = {'fill': PatternFill(start_color='ffe6f0', end_color='ffe6f0', fill_type='solid')}
                if txt == 'net discount':
                    style['font'] = Font(bold=True)
            
            if txt == 'grocery local [fcl]':
                blue_block = True
            if blue_block:
                style = {'fill': PatternFill(start_color='d6f0ff', end_color='d6f0ff', fill_type='solid')}
            if txt == 'drinks [fcd]':
                blue_block = False
            
            if txt == 'drinks [fcd] - alco':
                green1_block = True
            if green1_block:
                style = {'fill': PatternFill(start_color='e6ffe6', end_color='e6ffe6', fill_type='solid')}
            if txt == 'drinks [fcd] - non alco':
                green1_block = False
            
            if txt == 'add: opening inventory (alco)':
                green2_block = True
            if green2_block:
                style = {'fill': PatternFill(start_color='e6ffe6', end_color='e6ffe6', fill_type='solid')}
            if txt == 'add: closing inventory (non-alco)':
                green2_block = False
            
            if txt == 'bank charges/credit card charges':
                red_block = True
            if red_block:
                style = {'fill': PatternFill(start_color='ffe6e6', end_color='ffe6e6', fill_type='solid')}
            if txt == 'license fees':
                red_block = False
            
            # Specific row styling (overrides block colors)
            if txt == 'net sale':
                style = {'fill': PatternFill(start_color='e75480', end_color='e75480', fill_type='solid'), 'font': Font(bold=True)}
            elif txt == 'cost of food sold':
                style = {'font': Font(bold=True, underline='single')}
            elif txt == 'total food cost':
                style = {'fill': PatternFill(start_color='4f81bd', end_color='4f81bd', fill_type='solid'), 'font': Font(bold=True, color='FFFFFF')}
            elif txt in ['add: opening inventory (food)', 'less: closing inventory (food)']:
                style = {'fill': PatternFill(start_color='d6f0ff', end_color='d6f0ff', fill_type='solid')}
            elif txt == 'less: taxes (1/3rd)':
                style = {'fill': PatternFill(start_color='fffacd', end_color='fffacd', fill_type='solid'), 'font': Font(bold=True)}
            elif txt == 'net food cost':
                style = {'fill': PatternFill(start_color='4f81bd', end_color='4f81bd', fill_type='solid'), 'font': Font(bold=True, color='FFFFFF')}
            elif txt == 'disbursement':
                style = {'fill': PatternFill(start_color='4f81bd', end_color='4f81bd', fill_type='solid'), 'font': Font(bold=True, color='FFFFFF')}
            elif txt == 'cost of drinks sold':
                style = {'font': Font(bold=True, underline='single')}
            elif txt == 'total drinks cost' or txt == 'net drink cost':
                style = {'fill': PatternFill(start_color='5cb85c', end_color='5cb85c', fill_type='solid'), 'font': Font(bold=True, color='FFFFFF')}
            elif txt == 'gross profit':
                style = {'fill': PatternFill(start_color='d9534f', end_color='d9534f', fill_type='solid'), 'font': Font(bold=True, color='FFFFFF')}
            elif txt in ['expenses', 'expenses ']:
                style = {'font': Font(bold=True, underline='single')}
            elif txt == 'total non operating cost':
                style = {'fill': PatternFill(start_color='ff9900', end_color='ff9900', fill_type='solid'), 'font': Font(bold=True)}
            elif txt == 'net profit':
                style = {'fill': PatternFill(start_color='b30000', end_color='b30000', fill_type='solid'), 'font': Font(bold=True, color='FFFFFF')}
            # Apply style to the row
            if style:
                for c in range(1, ws.max_column+1):
                    if 'fill' in style:
                        ws.cell(row=r, column=c).fill = style['fill']
                    if 'font' in style:
                        ws.cell(row=r, column=c).font = style['font']
# Number formatting
# First, find the row numbers for NET PROFIT and Less: Taxes
net_profit_row = None
taxes_row = None
for r in range(2, ws.max_row+1):
    val = ws.cell(row=r, column=particulars_col).value
    if isinstance(val, str):
        txt = val.strip().lower()
        if txt == 'net profit':
            net_profit_row = r
        elif txt == 'less: taxes (1/3rd)':
            taxes_row = r
            break

for c in range(1, ws.max_column+1):
    col_name = ws.cell(row=1, column=c).value
    # Check if this is a percentage column
    is_percent_col = col_name and isinstance(col_name, str) and ('%' in col_name or col_name.strip().startswith('%') or col_name.strip().endswith('%'))
    
    for r in range(2, ws.max_row+1):
        try:
            val = ws.cell(row=r, column=c).value
            if isinstance(val, (int, float)) and val != 0:
                # Check if this row is between NET PROFIT and Less: Taxes - if so, use number format
                in_special_zone = net_profit_row and taxes_row and net_profit_row < r < taxes_row
                
                if is_percent_col and not in_special_zone:
                    # For percentage columns, format as percentage (except in special zone)
                    ws.cell(row=r, column=c).number_format = '0.00%'
                else:
                    # For regular number columns or special zone, use Indian number format
                    ws.cell(row=r, column=c).number_format = '#,##,##0'
        except Exception:
            pass
# Autosize columns
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    col_header = ws.cell(row=1, column=col[0].column).value
    
    # Check if this is a percentage column
    is_percent_col = col_header and isinstance(col_header, str) and ('%' in col_header or col_header.strip().startswith('%') or col_header.strip().endswith('%'))
    
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    
    # Set width: smaller for percentage columns, normal for others
    if is_percent_col:
        ws.column_dimensions[col_letter].width = min(max_length + 2, 10)  # Max 10 for % columns
    else:
        ws.column_dimensions[col_letter].width = max_length + 2
# Save to buffer
styled_buf = io.BytesIO()
wb.save(styled_buf)
styled_buf.seek(0)

# Get the latest month from the dataframe columns
latest_month = None
for col in df_to_show.columns:
    if col != 'PARTICULARS' and col != 'Branch':
        try:
            # Try to parse as date
            date_val = pd.to_datetime(col)
            if latest_month is None or date_val > latest_month:
                latest_month = date_val
        except:
            pass

# Format the latest month as mmm-yy
if latest_month:
    month_str = latest_month.strftime('%b-%y')
    file_name = f'P&L - Niko Foods LLP - {month_str}.xlsx'
else:
    file_name = 'P&L - Niko Foods LLP.xlsx'

# Create title row with download button and smaller remarks button
col1, col2, col3 = st.columns([3, 1, 1])
with col1:
    st.title("Niko Foods Profitability Dashboard")
with col2:
    st.write("")  # Add spacing
    if st.button('📝 Remarks', use_container_width=True, key='header_remarks_btn'):
        st.session_state.show_remarks_editor = not st.session_state.show_remarks_editor
with col3:
    st.write("")  # Add spacing
    st.download_button(
        label='📥 Download',
        data=styled_buf,
        file_name=file_name,
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        key=f'download_excel_{branch_option}',
        use_container_width=True
    )

# The table display remains below this logic

# Simple Remarks Management Section
if st.session_state.get('show_remarks_editor', False):
    st.markdown("---")
    st.subheader("📝 Remarks Management")

    col_a, col_b = st.columns(2)

    with col_a:
        st.markdown("**Add Remark to Cell**")
        # Get list of particulars for dropdown
        particulars_list = df_to_show['PARTICULARS'].tolist()
        selected_particular = st.selectbox(
            "Select Row (Particular)",
            options=particulars_list,
            key='remark_particular'
        )

        # Get month columns for dropdown
        month_cols_list = [col for col in df_to_show.columns if col != 'PARTICULARS']
        selected_column = st.selectbox(
            "Select Column (Month/Percentage)",
            options=month_cols_list,
            key='remark_column'
        )

        # Create unique cell key using formatted column names
        cell_key = f"{selected_particular}|{selected_column}"

        # Ensure the cell key uses the correct format for saving
        # Convert any raw datetime formats to the display format for consistency
        if '|' in cell_key:
            parts = cell_key.split('|')
            if len(parts) == 2:
                row_part, col_part = parts
                # Try to convert column part if it looks like a raw datetime
                try:
                    # Check if column part looks like a raw datetime (contains time component)
                    if ' ' in col_part and ':' in col_part:
                        # This is likely a raw datetime, convert it to formatted month
                        dt = pd.to_datetime(col_part)
                        formatted_col = dt.strftime('%b-%y')
                        cell_key = f"{row_part}|{formatted_col}"
                except Exception:
                    # If conversion fails, use the original cell_key
                    pass

        # Get existing remark if any
        existing_remark = st.session_state.remarks_data.get('remarks', {}).get(cell_key, '')

        remark_text = st.text_area(
            "Remark Text",
            value=existing_remark,
            height=100,
            key='remark_text',
            placeholder="Enter your remark here..."
        )

        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("💾 Save Remark", use_container_width=True):
                if remark_text.strip():
                    if 'remarks' not in st.session_state.remarks_data:
                        st.session_state.remarks_data['remarks'] = {}
                    st.session_state.remarks_data['remarks'][cell_key] = remark_text.strip()
                    if save_remarks(st.session_state.remarks_data):
                        st.success("✅ Remark saved and cell highlighted!")
                        st.rerun()
                else:
                    st.warning("Please enter a remark text.")

        with col_btn2:
            if st.button("🗑️ Delete Remark", use_container_width=True):
                if cell_key in st.session_state.remarks_data.get('remarks', {}):
                    del st.session_state.remarks_data['remarks'][cell_key]
                    if save_remarks(st.session_state.remarks_data):
                        st.success("✅ Remark deleted and highlight removed!")
                        st.rerun()
                else:
                    st.warning("No remark found for this cell.")

    with col_b:
        st.markdown("**Current Remarks**")
        if st.session_state.remarks_data.get('remarks'):
            for cell_key, remark in st.session_state.remarks_data['remarks'].items():
                particular, column = cell_key.split('|')
                st.markdown(f"📍 **{particular}** → **{column}**")
                st.caption(remark)
        else:
            st.info("No remarks added yet.")

    # Clear All Remarks button
    st.markdown("---")
    if st.button("🗑️ Clear All Remarks", use_container_width=True, type="secondary"):
        if st.session_state.remarks_data.get('remarks'):
            st.session_state.remarks_data['remarks'] = {}
            if save_remarks(st.session_state.remarks_data):
                st.success("✅ All remarks cleared!")
                st.rerun()
        else:
            st.info("No remarks to clear.")

# Determine unhidden columns for the selected branch using openpyxl
import openpyxl
excel_path = file_path
wb = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
ws = wb[branch_option]

# Get indexes of unhidden columns (1-based for openpyxl)
unhidden_col_indexes = [
    idx for idx, col in enumerate(ws.iter_cols(1, ws.max_column), 1)
    if not ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].hidden
]

# Get indexes of unhidden rows (1-based for openpyxl, skip header row)
unhidden_row_indexes = [
    idx - 2 for idx in range(2, ws.max_row + 1)  # -2 to convert to 0-based pandas index (row 2 in Excel = index 0 in pandas)
    if not ws.row_dimensions[idx].hidden
]

# Map to DataFrame column names (0-based for pandas)
df_cols = list(sheets_data_str[branch_option].columns)
unhidden_cols = [df_cols[idx-1] for idx in unhidden_col_indexes if idx-1 < len(df_cols)]

import numpy as np
import re
from datetime import datetime

df_to_show = sheets_data_str[branch_option][unhidden_cols].copy()
if unhidden_row_indexes:
    df_to_show = df_to_show.iloc[unhidden_row_indexes].reset_index(drop=True)

def indian_number_format(val):
    try:
        x = int(round(float(val)))
        s = str(x)[::-1]
        groups = []
        groups.append(s[:3])
        s = s[3:]
        while s:
            groups.append(s[:2])
            s = s[2:]
        return ','.join(groups)[::-1]
    except Exception:
        return ''

def format_percent(val):
    try:
        if pd.isnull(val) or val == '' or float(val) == 0:
            return ''
        return f"{float(val)*100:.2f}%"
    except Exception:
        return ''

def excel_month_fmt(col):
    # If column is datetime or looks like a month, format as 'Apr-25'
    if isinstance(col, datetime):
        return col.strftime('%b-%y')
    if isinstance(col, str):
        try:
            dt = pd.to_datetime(col)
            return dt.strftime('%b-%y')
        except Exception:
            return col
    return col

# Format headers
new_cols = []
for col in df_to_show.columns:
    new_cols.append(excel_month_fmt(col))
df_to_show.columns = new_cols

# Format values
skip_cols = []
for col in df_to_show.columns:
    if col.lower() == 'particulars' or col.lower() == 'branch' or col == '%' or str(col).strip().startswith('%') or str(col).strip().endswith('%'):
        skip_cols.append(col)

# Find index of 'Net Profit' row in 'PARTICULARS' column (case-insensitive)
net_profit_idx = None
for i, val in enumerate(df_to_show['PARTICULARS']):
    if isinstance(val, str) and val.strip().lower().startswith('net profit'):
        net_profit_idx = i
        break

# Apply formatting for rows before or at 'Net Profit'
for col in df_to_show.columns:
    if col in skip_cols:
        # Format % columns
        if col == '%' or str(col).strip().startswith('%') or str(col).strip().endswith('%'):
            df_to_show.loc[:net_profit_idx, col] = df_to_show.loc[:net_profit_idx, col].apply(format_percent)
        else:
            df_to_show.loc[:net_profit_idx, col] = df_to_show.loc[:net_profit_idx, col].replace([None, np.nan], '')
    else:
        # Format all other columns as Indian numbers
        df_to_show.loc[:net_profit_idx, col] = df_to_show.loc[:net_profit_idx, col].apply(lambda x: indian_number_format(x) if pd.notnull(x) and x != '' else '')

# For rows after Net Profit, do not move names; format numbers only
if net_profit_idx is not None:
    for i in range(net_profit_idx+1, len(df_to_show)):
        row = df_to_show.iloc[i]
        for col in df_to_show.columns:
            val = row[col]
            # Only format if it's a number
            try:
                # If value is numeric, format as Indian number
                if isinstance(val, (int, float)) and not pd.isnull(val):
                    df_to_show.at[i, col] = indian_number_format(val)
                # If value is string but represents a number
                elif isinstance(val, str) and val.replace(',', '').replace('.', '').isdigit():
                    df_to_show.at[i, col] = indian_number_format(val)
                # Else leave as-is (expense name or blank)
            except Exception:
                pass

# Ensure all None and np.nan are shown as blanks
df_to_show = df_to_show.replace([None, np.nan], '')

def highlight_sales_block(df):
    sales_start = "SALES"
    sales_end = "TOTAL SALES AND SERVICE CHARGES"
    pink_rows = ["LESS: DISCOUNT", "LESS: ADJUSTED ( NET OF GST)", "NET DISCOUNT"]
    highlight = False
    highlights = []
    
    # Get remarks data
    remarks_data = st.session_state.remarks_data
    user_remarks = remarks_data.get('remarks', {})
    
    for row_idx, row in df.iterrows():
        row_styles = ['' for _ in row]
        if isinstance(row['PARTICULARS'], str):
            particulars = row['PARTICULARS'].strip().lower()
            # Deep pink Net Sale row
            if particulars == "net sale":
                row_styles = ['background-color: #e75480; font-weight: bold' for _ in row]
            # Bold and underline COST OF FOOD SOLD
            elif particulars == "cost of food sold":
                row_styles = ['font-weight: bold; text-decoration: underline' for _ in row]
            # Deeper blue and bold TOTAL FOOD COST
            elif particulars == "total food cost":
                row_styles = ['background-color: #4f81bd; color: white; font-weight: bold' for _ in row]
            # Light blue for inventory rows
            elif particulars in ["add: opening inventory", "less: closing inventory"]:
                row_styles = ['background-color: #d6f0ff' for _ in row]
            # Light yellow for taxes
            elif particulars == "less: taxes (1/3rd)":
                row_styles = ['background-color: #fffacd; font-weight: bold' for _ in row]
            # Deeper blue and bold NET FOOD COST
            elif particulars == "net food cost":
                row_styles = ['background-color: #4f81bd; color: white; font-weight: bold' for _ in row]
            # Deep blue and bold DISBURSEMENT
            elif particulars == "disbursement":
                row_styles = ['background-color: #4f81bd; color: white; font-weight: bold' for _ in row]
            else:
                # Light blue block: GROCERY [FCL] to DRINKS [FCD]
                blue_start = "grocery [fcl]"
                blue_end = "drinks [fcd]"
                if not hasattr(highlight_sales_block, 'in_blue_block'):
                    highlight_sales_block.in_blue_block = False
                if blue_start == particulars:
                    highlight_sales_block.in_blue_block = True
                if highlight_sales_block.in_blue_block:
                    row_styles = ['background-color: #d6f0ff' for _ in row]
                if blue_end == particulars:
                    highlight_sales_block.in_blue_block = False

                # Light green block: DRINKS [FCD] - ALCO to DRINKS [FCD] - NON ALCO
                green1_start = "drinks [fcd] - alco"
                green1_end = "drinks [fcd] - non alco"
                if not hasattr(highlight_sales_block, 'in_green1_block'):
                    highlight_sales_block.in_green1_block = False
                if green1_start == particulars:
                    highlight_sales_block.in_green1_block = True
                if highlight_sales_block.in_green1_block:
                    row_styles = ['background-color: #e6ffe6' for _ in row]
                if green1_end == particulars:
                    highlight_sales_block.in_green1_block = False

                # Light green block: ADD: OPENING INVENTORY (ALCO) to ADD: CLOSING INVENTORY (NON-ALCO)
                green2_start = "add: opening inventory (alco)"
                green2_end = "add: closing inventory (non-alco)"
                if not hasattr(highlight_sales_block, 'in_green2_block'):
                    highlight_sales_block.in_green2_block = False
                if green2_start == particulars:
                    highlight_sales_block.in_green2_block = True
                if highlight_sales_block.in_green2_block:
                    row_styles = ['background-color: #e6ffe6' for _ in row]
                if green2_end == particulars:
                    highlight_sales_block.in_green2_block = False

                # Bold and underline COST OF DRINKS SOLD
                if particulars == "cost of drinks sold":
                    row_styles = ['font-weight: bold; text-decoration: underline' for _ in row]
                # Deeper green and bold TOTAL DRINKS COST
                elif particulars == "total drinks cost":
                    row_styles = ['background-color: #5cb85c; color: white; font-weight: bold' for _ in row]
                # Deeper green and bold NET DRINK COST
                elif particulars == "net drink cost":
                    row_styles = ['background-color: #5cb85c; color: white; font-weight: bold' for _ in row]
                # Deep red and bold Gross Profit
                elif particulars == "gross profit":
                    row_styles = ['background-color: #d9534f; color: white; font-weight: bold' for _ in row]
                # Bold and underline Expenses
                elif particulars == "expenses":
                    row_styles = ['font-weight: bold; text-decoration: underline' for _ in row]
                # Light red block: BANK CHARGES/CREDIT CARD CHARGES to LICENSE FEES
                else:
                    red_start = "bank charges/credit card charges"
                    red_end = "license fees"
                    if not hasattr(highlight_sales_block, 'in_red_block'):
                        highlight_sales_block.in_red_block = False
                    if red_start == particulars:
                        highlight_sales_block.in_red_block = True
                    if highlight_sales_block.in_red_block:
                        row_styles = ['background-color: #ffe6e6' for _ in row]
                    if red_end == particulars:
                        highlight_sales_block.in_red_block = False
                    # Bold and orange TOTAL NON OPERATING COST
                    if particulars == "total non operating cost":
                        row_styles = ['background-color: #ff9900; font-weight: bold' for _ in row]
                    # Bold and deep red NET PROFIT
                    elif particulars == "net profit":
                        row_styles = ['background-color: #b30000; color: white; font-weight: bold' for _ in row]

                # Pink block
                for pink in pink_rows:
                    if pink.lower() == particulars:
                        if pink == "NET DISCOUNT":
                            row_styles = ['background-color: #ffe6f0; font-weight: bold' for _ in row]
                        else:
                            row_styles = ['background-color: #ffe6f0' for _ in row]
                        break
                else:  # Only check sales block if not pink or net sale or cost of food sold or blue/green/red blocks
                    if sales_start.lower() in particulars:
                        highlight = True
                    if highlight:
                        if sales_end.lower() in particulars:
                            row_styles = ['background-color: #ffe066; font-weight: bold' for _ in row]
                            highlight = False
                        else:
                            row_styles = ['background-color: #fff9c4' for _ in row]
        
        # Apply automatic highlighting for cells with remarks (overrides default styling)
        particular_val = row['PARTICULARS']
        for col_idx, col_name in enumerate(df.columns):
            cell_key = f"{particular_val}|{col_name}"
            
            # Check if cell has a remark - automatically highlight it (overrides all other styling)
            if cell_key in user_remarks:
                # Automatic highlight for cells with remarks (light orange/peach color with important flags)
                row_styles[col_idx] = 'background-color: #FFE4B5 !important; font-weight: bold !important; border: 3px solid #ff6b6b !important; position: relative'
        
        highlights.append(row_styles)
    return pd.DataFrame(highlights, columns=df.columns)

def style_table(df):
    # Identify month columns (columns that look like 'Jul-25', 'Aug-25', etc.)
    month_pattern = r'^[A-Z][a-z]{2}-\d{2}$'
    month_cols = [col for col in df.columns if re.match(month_pattern, str(col))]
    
    # Build table styles
    table_styles = [
        {
            'selector': 'th',
            'props': [
                ('background-color', '#003366'),
                ('color', 'white'),
                ('font-weight', 'bold'),
                ('font-size', '16px'),
                ('text-align', 'center')
            ]
        },
        {
            'selector': 'td',
            'props': [
                ('text-align', 'center')
            ]
        }
    ]
    
    # Add specific width for month columns
    for idx, col in enumerate(df.columns):
        if col in month_cols:
            table_styles.append({
                'selector': f'th:nth-child({idx+1}), td:nth-child({idx+1})',
                'props': [
                    ('min-width', '120px'),
                    ('max-width', '120px'),
                    ('width', '120px'),
                    ('text-align', 'center')
                ]
            })
        elif col == 'PARTICULARS':
            table_styles.append({
                'selector': f'th:nth-child({idx+1}), td:nth-child({idx+1})',
                'props': [
                    ('text-align', 'left'),
                    ('min-width', '250px')
                ]
            })
        elif '%' in str(col):
            table_styles.append({
                'selector': f'th:nth-child({idx+1}), td:nth-child({idx+1})',
                'props': [
                    ('min-width', '80px'),
                    ('max-width', '80px'),
                    ('width', '80px'),
                    ('text-align', 'center')
                ]
            })
    
    styler = df.style.set_table_styles(table_styles).hide(axis='index')
    return styler.apply(highlight_sales_block, axis=None)

# Display the table with remarks tooltip info
table_html = style_table(df_to_show).to_html(escape=False)

# Add tooltip/hover info for cells with remarks
if st.session_state.remarks_data.get('remarks'):
    # Get the remarks data for tooltip creation
    remarks_data = st.session_state.remarks_data.get('remarks', {})

    # Parse the HTML and add data-remark attributes to cells with remarks
    soup = BeautifulSoup(table_html, 'html.parser')
    table = soup.find('table')

    if table:
        # Find all rows in the table body (skip header)
        rows = table.find_all('tr')[1:]  # Skip header row

        for row_idx, row in enumerate(rows):
            # Get the particulars value from the first cell (row header)
            first_cell = row.find('td') or row.find('th')
            if first_cell:
                particular_text = first_cell.get_text(strip=True)
                if particular_text == 'PARTICULARS':
                    continue  # Skip the header row

                # Find all data cells in this row
                data_cells = row.find_all('td')[1:]  # Skip the particulars cell

                for col_idx, cell in enumerate(data_cells):
                    # Get the column name from the header
                    header_row = table.find('tr')
                    if header_row:
                        headers = header_row.find_all(['th', 'td'])
                        if col_idx + 1 < len(headers):  # +1 because we skip the particulars column
                            col_name = headers[col_idx + 1].get_text(strip=True)
                            cell_key = f"{particular_text}|{col_name}"

                            if cell_key in remarks_data:
                                # Add the remark as a data attribute
                                remark_text = remarks_data[cell_key]
                                cell['data-remark'] = remark_text
                                # Also add title attribute as fallback
                                cell['title'] = remark_text

    # Convert back to string
    table_html = str(soup)

    st.markdown("""
    <style>
    /* Tooltip styling for cells with remarks */
    .freeze-header-table-container td[data-remark] {
        position: relative !important;
        cursor: help !important;
    }

    .freeze-header-table-container td[data-remark]:hover::after {
        content: attr(data-remark);
        position: absolute !important;
        background: rgba(51, 51, 51, 0.95) !important;
        color: white !important;
        padding: 10px 15px !important;
        border-radius: 8px !important;
        font-size: 13px !important;
        font-family: Arial, sans-serif !important;
        white-space: pre-wrap !important;
        max-width: 350px !important;
        min-width: 200px !important;
        z-index: 1000 !important;
        bottom: 100% !important;
        left: 50% !important;
        transform: translateX(-50%) !important;
        margin-bottom: 8px !important;
        box-shadow: 0 4px 12px rgba(0,0,0,0.4) !important;
        pointer-events: none !important;
        line-height: 1.4 !important;
    }

    /* Add a small indicator that this cell has a remark */
    .freeze-header-table-container td[data-remark]::before {
        content: "💬";
        position: absolute !important;
        top: 2px !important;
        right: 2px !important;
        font-size: 10px !important;
        opacity: 0.7 !important;
        z-index: 1 !important;
    }
    </style>
    """, unsafe_allow_html=True)

st.markdown(
    f'<div class="freeze-header-table-container">{table_html}</div>',
    unsafe_allow_html=True
)

# Summary Reports & Charts
st.markdown("---")
st.header("Summary Reports & Charts")

# KPI Tiles for latest month (selected branch only)
branch_df = sheets_data_str[branch_option]

# Identify month columns (exclude non-month columns)
non_month_cols = ['PARTICULARS', 'Branch', 'Month']
month_cols = [col for col in branch_df.columns if col not in non_month_cols]

# Try to parse columns as dates and find the latest
month_col_dates = []
for col in month_cols:
    try:
        dt = pd.to_datetime(col, format='%b-%y', errors='coerce')
        if not pd.isnull(dt):
            month_col_dates.append((col, dt))
    except Exception:
        continue
if month_col_dates:
    # Sort and pick the latest
    month_col_dates.sort(key=lambda x: x[1])
    latest_month_col, latest_month_dt = month_col_dates[-1]
    # KPIs and their row labels
    kpi_labels = [
        ('Total Sales and Service Charge', 'TOTAL SALES AND SERVICE CHARGES'),
        ('Net Food Cost', 'NET FOOD COST'),
        ('Net Drink Cost', 'NET DRINK COST'),
        ('Gross Profit', 'GROSS PROFIT'),
        ('Total Non Operating Cost', 'TOTAL NON OPERATING COST'),
        ('Net Profit', 'NET PROFIT'),
    ]
    kpi_results = []
    for kpi_name, row_label in kpi_labels:
        row = branch_df[branch_df['PARTICULARS'].str.strip().str.upper() == row_label]
        if not row.empty and latest_month_col in row.columns:
            value = row[latest_month_col].values[0]
            if value is not None and value != '' and value != 0:
                try:
                    value_fmt = indian_number_format(value)
                except Exception:
                    value_fmt = str(value)
            else:
                value_fmt = '-'
        else:
            value_fmt = '-'
        kpi_results.append((kpi_name, value_fmt))
    # Show tiles: 3 per row, smaller, colored
    st.markdown(f'#### Latest Month KPIs ({latest_month_col})')
    colors = [
        '#e3f2fd', '#fff9c4', '#ffe0b2', '#c8e6c9', '#f8bbd0', '#d1c4e9'
    ]
    tile_html = """
    <style>
    .kpi-row {{ display: flex; flex-wrap: wrap; gap: 1rem; margin-bottom: 1rem; }}
    .kpi-tile {{
        flex: 1 1 calc(33% - 1rem);
        min-width: 180px;
        background: {bg};
        border-radius: 12px;
        padding: 0.7rem 0.5rem 0.5rem 0.5rem;
        box-shadow: 0 2px 8px rgba(0,0,0,0.04);
        text-align: center;
        margin-bottom: 0.5rem;
    }}
    .kpi-label {{ font-size: 1rem; color: #333; margin-bottom: 0.2rem; font-weight: 600; }}
    .kpi-value {{ font-size: 1.5rem; color: #003366; font-weight: bold; letter-spacing: 1px; }}
    @media (max-width: 800px) {{
        .kpi-tile {{ flex: 1 1 100%; min-width: 140px; }}
    }}
    </style>
    <div class="kpi-row">
    {tiles}
    </div>
    """
    tiles = ""
    for idx, (kpi_name, value_fmt) in enumerate(kpi_results):
        bg = colors[idx % len(colors)]
        tiles += f'<div class="kpi-tile" style="background:{bg}"><div class="kpi-label">{kpi_name}</div><div class="kpi-value">{value_fmt}</div></div>'
    st.markdown(tile_html.format(tiles=tiles, bg='{bg}'), unsafe_allow_html=True)

# Niko Sales Trend
if 'Sale' in full_df.columns and 'Month' in full_df.columns:
    fig_sales = px.line(full_df, x='Month', y='Sale', title="Niko Monthly Sales Trend")
    st.plotly_chart(fig_sales, use_container_width=True)

# Niko Profit Trend
if 'Month' in full_df.columns and 'Profit' in full_df.columns:
    fig_profit = px.line(full_df, x='Month', y='Profit', title="Niko Monthly Profit Trend")
    st.plotly_chart(fig_profit, use_container_width=True)
